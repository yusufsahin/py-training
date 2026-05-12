from openpyxl import Workbook, load_workbook
from pathlib import Path

current_dir = Path(__file__).parent
file_path = current_dir / "ogrenciler.xlsx"


# Excel dosyası oluştur
def excel_olustur():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ogrenciler"

    # Sütun başlıkları
    ws.append(["id", "isim", "soyisim", "bolum"])

    wb.save(file_path)


# Excel'e öğrenci ekle
def ogrenci_ekle(id, isim, soyisim, bolum):
    wb = load_workbook(file_path)
    ws = wb["Ogrenciler"]

    # Yeni satır ekle
    ws.append([id, isim, soyisim, bolum])

    wb.save(file_path)


# Excel'deki tüm öğrencileri oku
def ogrencileri_listele():
    wb = load_workbook(file_path)
    ws = wb["Ogrenciler"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)


# İsme göre ara
def ogrenci_ara(aranan_isim):
    wb = load_workbook(file_path)
    ws = wb["Ogrenciler"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        id, isim, soyisim, bolum = row

        if isim.lower() == aranan_isim.lower():
            print("Bulundu:", row)


# Öğrenci güncelle
def ogrenci_guncelle(ogrenci_id, yeni_bolum):
    wb = load_workbook(file_path)
    ws = wb["Ogrenciler"]

    for row_no in range(2, ws.max_row + 1):
        id = ws.cell(row=row_no, column=1).value

        if id == ogrenci_id:
            ws.cell(row=row_no, column=4).value = yeni_bolum
            wb.save(file_path)
            print("Öğrenci güncellendi")
            return

    print("Öğrenci bulunamadı")


# Öğrenci sil
def ogrenci_sil(ogrenci_id):
    wb = load_workbook(file_path)
    ws = wb["Ogrenciler"]

    for row_no in range(2, ws.max_row + 1):
        id = ws.cell(row=row_no, column=1).value

        if id == ogrenci_id:
            ws.delete_rows(row_no)
            wb.save(file_path)
            print("Öğrenci silindi")
            return

    print("Öğrenci bulunamadı")


# Kullanım
excel_olustur()

ogrenci_ekle(1, "John", "Doe", "Physics")
ogrenci_ekle(2, "Sam", "Doe", "Math")
ogrenci_ekle(3, "Alice", "Brown", "Chemistry")

print("Tüm öğrenciler:")
ogrencileri_listele()

print("Arama sonucu:")
ogrenci_ara("Sam")

ogrenci_guncelle(2, "Computer Science")

print("Güncellemeden sonra:")
ogrencileri_listele()

ogrenci_sil(1)

print("Silmeden sonra:")
ogrencileri_listele()